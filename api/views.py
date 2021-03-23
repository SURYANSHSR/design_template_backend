from django.http import JsonResponse
import sqlite3
import pandas as pd
import json
import openpyxl


def allrecords1(request):
    con = sqlite3.connect('./tyredata.sqlite')
    cursorObj = con.cursor()
    cursorObj.execute("create table if not exists tes (id TEXT primary key not null, product_code TEXT, revision TEXT, tyre_size TEXT, tyre_pattern TEXT);")
    con.commit()
    cursorObj.execute("SELECT * FROM 'tes' ;")
    print(123)
    rows = cursorObj.fetchall()
    rs = pd.DataFrame(data=rows, columns=["id", "product_code", "revision", "tyre_size", "tyre_pattern"])
    # print(rs)
    js = rs.to_json()
    ps = json.loads(js)

    return JsonResponse(ps)


def check(request):
    data={
      "x":"123",
      "y":"234"
    }


    return JsonResponse(data)


def recorddetails(request):
    id = request.GET.get('q', '')
    con = sqlite3.connect('./tyredata.sqlite')
    cursorObj = con.cursor()
    cursorObj.execute("create table if not exists tes (id TEXT primary key not null, product_code TEXT, revision TEXT, tyre_size TEXT, tyre_pattern TEXT);")
    query="SELECT * FROM 'tes' where 'id'='"+id+"';"
    print(query)
    cursorObj.execute('SELECT * FROM "tes" where "id"="'+id+'";')
    print(123)
    rows = cursorObj.fetchall()
    rs = pd.DataFrame(data=rows, columns=["id", "product_code", "revision", "tyre_size", "tyre_pattern"])
    # print(rs) 
    js = rs.to_json()
    ps = json.loads(js)

    return JsonResponse(ps)





def createrecord(request):
    data = json.loads(request.POST.get('data'))
    product_code = data["productCode"]
    revision=data["revision"]
    tyre_size=data["tyreSize"]
    tyre_pattern=data["tyrePattern"]
    id=product_code+revision
    values = [id,product_code,revision,tyre_size,tyre_pattern ];
    con = sqlite3.connect('./tyredata.sqlite')
    cursorObj = con.cursor()
    cursorObj.execute("create table if not exists tes (id TEXT primary key not null, product_code TEXT, revision TEXT, tyre_size TEXT, tyre_pattern TEXT);")
    cursorObj.execute('insert or replace into tes ("id","product_code","revision", "tyre_size", "tyre_pattern") values (?,?,?,?,?)',values)
    con.commit()
    query="SELECT * FROM 'tes' where 'id'='"+id+"';"
    print(query)
    cursorObj.execute('SELECT * FROM "tes" where "id"="'+id+'";')
    print(123)
    rows = cursorObj.fetchall()
    rs = pd.DataFrame(data=rows, columns=["id", "product_code", "revision", "tyre_size", "tyre_pattern"])
    # print(rs) 
    js = rs.to_json()
    ps = json.loads(js)

    return JsonResponse(ps)

def updaterecord(request):
    data = json.loads(request.POST.get('data'))
    product_code = data["productCode"]
    revision=data["revision"]
    tyre_size=data["tyreSize"]
    tyre_pattern=data["tyrePattern"]
    id=product_code+revision
    values = [id,product_code,revision,tyre_size,tyre_pattern ];
    con = sqlite3.connect('./tyredata.sqlite')
    cursorObj = con.cursor()
    cursorObj.execute("create table if not exists tes (id TEXT primary key not null, product_code TEXT, revision TEXT, tyre_size TEXT, tyre_pattern TEXT);")
    cursorObj.execute('insert or replace into tes ("id","product_code","revision", "tyre_size", "tyre_pattern") values (?,?,?,?,?)',values)
    con.commit()
    query="SELECT * FROM 'tes' where 'id'='"+id+"';"
    print(query)
    cursorObj.execute('SELECT * FROM "tes" where "id"="'+id+'";')
    print(123)
    rows = cursorObj.fetchall()
    rs = pd.DataFrame(data=rows, columns=["id", "product_code", "revision", "tyre_size", "tyre_pattern"])
    # print(rs) 
    js = rs.to_json()
    ps = json.loads(js)

    return JsonResponse(ps)

from django.http import FileResponse

def export_excel(request):
   id = request.GET.get('q', '')
   con = sqlite3.connect('./tyredata.sqlite')
   cursorObj = con.cursor()
   cursorObj.execute("create table if not exists tes (id TEXT primary key not null, product_code TEXT, revision TEXT, tyre_size TEXT, tyre_pattern TEXT);")
   cursorObj.execute('SELECT * FROM "tes" where "id"="'+id+'";')
   print(123)
   rows = cursorObj.fetchall()
   print(rows[0][1])
   rs = pd.DataFrame(data=rows, columns=["id", "product_code", "revision", "tyre_size", "tyre_pattern"])
   # print(rs) 
   js = rs.to_json()
   ps = json.loads(js)

   wb=openpyxl.load_workbook('./Sample.xlsx')
   sh1= wb['Sheet1']
   sh1.cell(row=1,column=2,value=rows[0][1])
   sh1.cell(row=2,column=2,value=rows[0][2])
   sh1.cell(row=3,column=2,value=rows[0][3])
   sh1.cell(row=4,column=2,value=rows[0][4])
   wb.save('./Sample1.xlsx')
   exl = open('./Sample1.xlsx', 'rb')

   response = FileResponse(exl)

   return response